import os
import dotenv
import json
import datetime as dt
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from collections import defaultdict

"""Program to document toggle timesheets onto a given spreadsheet.

Will only work if timesheet contains only 2 entries per day.
"""

COL_DATE = "A"
COL_MORNING_IN = "B"
COL_MORNING_OUT = "C"
COL_AFTERNOON_IN = "D"
COL_AFTERNOON_OUT = "E"
START_ROW = 11

DEBUG = False

dotenv.load_dotenv()
workspace_id = os.getenv("WORKSPACE_ID")
user_agent = os.getenv("TOGGL_EMAIL")
api_token = os.getenv("TOGGL_API_TOKEN")
excel_path = os.getenv("EXCEL_PATH")

def find_last_filled_row(ws) -> int:
    """Find last row filled based on if morning is populated."""
    row = START_ROW
    while (ws[f"{COL_MORNING_IN}{row}"].value and ws[f"{COL_MORNING_OUT}{row}"].value
        or (ws[f"{COL_AFTERNOON_IN}{row}"].value and ws[f"{COL_AFTERNOON_OUT}{row}"].value)):
        if ws[f"{COL_MORNING_IN}{row}"].value == "Weekend":
            row += 2
        else:
            row += 1
    return row - 1

def get_fetch_range(ws, last_row:int) -> tuple[dt.datetime]:
    """Find range of sheet to fill."""
    sheet_start_date = ws[f"{COL_DATE}{START_ROW}"].value

    if last_row < START_ROW:
        start_date = sheet_start_date
    else:
        day_count = last_row - START_ROW
        start_date = sheet_start_date + dt.timedelta(days=day_count)
        
    end_date = dt.datetime.today()
    return start_date, end_date

def fetch_toggl_entries(start_date:dt.date,end_date:dt.date) -> list:
    """API call to toggl to get required information."""
    url = r"https://api.track.toggl.com/reports/api/v2/details"
    params = {
        "workspace_id": workspace_id,
        "since": start_date.date().isoformat(),
        "until": end_date.date().isoformat(),
        "user_agent": user_agent,
        "page": 1
    }

    entries = []

    while True:
        r = requests.get(
            url,
            params=params,
            auth=HTTPBasicAuth(api_token,"api_token")
        )
        r.raise_for_status()
        data = r.json()

        entries.extend(data["data"])

        if params["page"] * data["per_page"] >= data["total_count"]:
            break
        params["page"] +=1

    if DEBUG:
        json_fp = "./api_output.json"
        with open(json_fp,'w') as wf:
            json.dump(data,wf)
        print(f"Exported API response to {json_fp}.")
    
    return entries

def group_entries_by_date(entries:list) -> dt.datetime:
    """Group entries by date."""
    days = defaultdict(list)

    for e in entries:
        start = dt.datetime.fromisoformat(e["start"])
        end = dt.datetime.fromisoformat(e["end"])
        days[start.date()].append((start,end))

    return days

def write_times(ws, start_row:int, grouped_entries:dt.datetime):
    """Write timesheet times to sheet."""
    row = start_row

    for day in sorted(grouped_entries.keys()):
        blocks = sorted(grouped_entries[day], key=lambda x: x[0])

        if len(blocks) != 2:
            raise ValueError(f"{day} does not have exactly two time blocks")
    
        (m_in, m_out), (a_in, a_out) = blocks

        m_out += dt.timedelta(minutes=2)
        a_out += dt.timedelta(minutes=2)
        
        if DEBUG:
            print("BLOCKS = ",blocks)
            print("COL_MORNING_IN = ",ws[f"{COL_MORNING_IN}{row}"])
            print("COL_MORNING_OUT = ",ws[f"{COL_MORNING_IN}{row}"])
            print("COL_AFTERNOON_IN = ",ws[f"{COL_MORNING_IN}{row}"])
            print("COL_AFTERNOON_OUT = ",ws[f"{COL_MORNING_IN}{row}"])

        selected_date = ws[f"{COL_DATE}{row}"].value.date()
        
        while selected_date != m_in.date() or selected_date != a_in.date():
            row += 1
            selected_date = ws[f"{COL_DATE}{row}"].value.date()

        ws[f"{COL_MORNING_IN}{row}"] = m_in.strftime("%H:%M")
        ws[f"{COL_MORNING_OUT}{row}"] = m_out.strftime("%H:%M")
        ws[f"{COL_AFTERNOON_IN}{row}"] = a_in.strftime("%H:%M")
        ws[f"{COL_AFTERNOON_OUT}{row}"] = a_out.strftime("%H:%M")

        row +=1


def main():
    wb = load_workbook(excel_path)
    print("Workbook loaded.")
    ws = wb.active

    last_row = find_last_filled_row(ws)
    start_date, end_date = get_fetch_range(ws, last_row)

    if start_date > end_date:
        print("Nothing to fill.")
        return
    
    entries = fetch_toggl_entries(start_date,end_date)
    grouped = group_entries_by_date(entries)

    write_times(ws, last_row + 1, grouped)

    wb.save(excel_path)

    print("Timesheet updated successfully.")

if __name__ == "__main__":
    main()